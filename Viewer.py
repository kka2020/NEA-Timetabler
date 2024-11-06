import openpyxl
import StructTools

class Viewer():
    def __init__(self, spec):
        self.spec = spec
        
    def viewTimetable(self, timetable, course_to_room_by_session, out_doc):
        """
        Produces an Excel spreadsheet that displays the timetables
        for each course and each lecturer in their own sheets, in 
        the form of a table with days as columns and periods as rows.
        
        Args:
            timetable (dict): The timetable data structure outputted from
                            from the timetabling component.
        """
        # Creates spreadsheet object
        workbook = openpyxl.Workbook()
        
        del workbook['Sheet']
        # Creates a sheet for each course and populates it with its timetable
        for course in self.spec.courses:
            workbook.create_sheet(course)
            self.populate(workbook, timetable, course_to_room_by_session, course=course)
        
        # Creates a sheet for each lecturer and populates it with their timetable
        for lecturer in range(self.spec.lecturer_count):
            workbook.create_sheet(f"Lecturer {self.spec.lecturer_names[lecturer]}")
            self.populate(workbook, timetable, course_to_room_by_session, lecturer=lecturer)
        
        for room in StructTools.chain(*[rooms for rooms in self.spec.rooms.values()]):
            workbook.create_sheet(f"Room {room}")
            self.populate(workbook, timetable, course_to_room_by_session, room=room)
        
        # Saves spreadsheet
        # User can then view it in Excel
        workbook.save(filename=out_doc)

    def populate(self, workbook, timetable, course_to_room_by_session, course=None, lecturer=None, room=None):
        """
        Populates a sheet with the particular course's or lecturer's
        timetable

        Args:
            workbook (Workbook): The spreadsheet we are writing to
            timetable (dict): Timetable data structure
            course (str, optional): The name of the course (if the current
                                    entity is a course). Defaults to None.
            lecturer (int, optional): The ID of the lecturer. Defaults to None.
        """
        # If lecturer is None, then we know we are populating for a course
        # If otherwise, then we know we are populating for a lecturer
        if course is not None:
            sheet = workbook[course]
        elif lecturer is not None:
            sheet = workbook[f"Lecturer {self.spec.lecturer_names[lecturer]}"]
        else:
            sheet = workbook[f"Room {room}"]
        
        # Generates row headers
        for period in range(self.spec.periods):
            sheet[f"A{2 + period}"] = f"P{period + 1}"
        
        # Generates column headers
        for day in range(self.spec.days):
            sheet[chr(66 + day) + "1"] = f"D{day + 1}"
        
        # Maps the short-codes for each session type to their full names
        full_session_type = {"lec" : "lecture", "sem" : "seminar", "lab" : "lab"}
        
        # Iterates throug each period
        for (period, sessions) in timetable.items():
            # Calculates the cell corresponding to the current period
            # using the number of periods per day
            cell = chr(66 + period // self.spec.periods) + str(2 + period % self.spec.periods)
            
            session_info = ""
            
            for session in sessions: 
                # Extracts module from session string identifier
                mod = int(session[0][4:6]) if session[0][5] != "," else int(session[0][4])
                
                # Handles course-table info
                if (course is not None) and (mod in self.spec.course_to_modules[course]):
                    # Adds information strictly from the session's string identifier
                    # (i.e. module name and session type)
                    session_info = f"Module {self.spec.module_names[mod]} {full_session_type[session[0][:3]]}; "
                    
                    # Adds room specific information for each room that the current course is in
                    # (i.e. the name of the given room, the number of students from the course in there and which lecturer is teaching in there)
                    for (curr_room, student_count) in course_to_room_by_session[session[0]][course].items():
                        room_lecturer = session[2][session[1].index(curr_room)]
                        session_info += f"{student_count} students in room {curr_room} with lecturer {self.spec.lecturer_names[room_lecturer]}; "
                # Handles lecturer-table info
                elif (lecturer is not None) and (lecturer in session[2]):
                    # Gets the room that the lecturer is teaching in for this period
                    lecturer_room = session[1][session[2].index(lecturer)]
                    
                    # Adds information from session's string identifier + the current room
                    session_info = f"Module {self.spec.module_names[mod]} {full_session_type[session[0][:3]]} in {lecturer_room}; "

                    # Adds information about the number of students in the room and from which courses they're from
                    for (curr_course, rooms_to_students) in course_to_room_by_session[session[0]].items():
                        if lecturer_room in rooms_to_students:
                            session_info += f"{rooms_to_students[lecturer_room]} students from {curr_course} course; "
                # Handles room-table info
                elif (room is not None) and (room in session[1]):
                    # Gets the lecturer teaching the the given room
                    room_lecturer = session[2][session[1].index(room)]
                    
                    # Adds information from session's string identifier + the current lecturer
                    session_info = f"Module {self.spec.module_names[mod]} {full_session_type[session[0][:3]]} session lead by lecturer {self.spec.lecturer_names[room_lecturer]}; "
                    
                    # Adds information about the number of students in the room and from which courses they're from
                    for (curr_course, rooms_to_students) in course_to_room_by_session[session[0]].items():
                        if room in rooms_to_students:
                            session_info += f"{rooms_to_students[room]} students from {curr_course} course; "
                
                # Fills cell in with the current session's info string
                sheet[cell] = session_info